import streamlit as st
import ui as ui
import data_processing as dp 
import pandas as pd
import numpy as np
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles.fills import PatternFill
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import tempfile

def form_abertura(content, temp_dir):
    df_abertura = content

    # Arrumar nomes das colunas
    df_abertura.columns = df_abertura.columns.str.strip().str.lower().str.replace(' ', '_')

    # Padronizando a escrita da coluna 'abertura/ampliação'
    df_abertura['abertura_/_ampliação'] = df_abertura['abertura_/_ampliação'].str.upper()

    # Filtrando a última data e o tipo de formulário
    df_filtrado_abertura = df_abertura[df_abertura['abertura_/_ampliação'] == 'ABERTURA']
    latest_date = df_filtrado_abertura['data_do_mapeamento'].max()
    df_filtrado_abertura = df_filtrado_abertura[df_filtrado_abertura['data_do_mapeamento'] == latest_date]

    # Removendo o DataFrame original para liberar memória
    del df_abertura

    if df_filtrado_abertura.empty:
        print("\nNão há solicitação de abertura.")
    else:
        solicitantes = df_filtrado_abertura['coletor_escritório_(responsável)'].unique()
    
        for solicitante in solicitantes:
            filtro_solicitante = df_filtrado_abertura[df_filtrado_abertura['coletor_escritório_(responsável)'] == solicitante]
            ufs_escritorio = filtro_solicitante['uf_do_escritório'].unique()
        
            for uf_escritorio in ufs_escritorio:
                filtro_uf = filtro_solicitante[filtro_solicitante['uf_do_escritório'] == uf_escritorio]
                jobs = filtro_uf['job'].unique()
            
                for job in jobs:
                    wb_abertura = load_workbook(filename="scr/template_abertura.xlsx")
                    form = filtro_uf[filtro_uf['job'] == job]

                    # Inicializando listas para armazenar os contatos separados
                    telefone_final = []
                    email_final = []
                    site_final = []
                    end_final = []
                    observacao_final = []

                    form['meio_de_contato'].fillna("-", inplace=True)

                    for meio_contato in form['meio_de_contato'].to_list():
                        meio_contato_aux = re.split(r" / |/", meio_contato)
                        meio_contato_aux = [contato.strip() for contato in meio_contato_aux]
    
                        # Inicializando listas temporárias para cada tipo de contato
                        telefone_temp = []
                        email_temp = []
                        site_temp = []
                        endereco_temp = []
                        observacao_temp = []
    
                        # Separando os contatos
                        for contato in meio_contato_aux:
                            # Verificando se é um telefone
                            if re.match(r"\(\d{2}\)\s?\d{4,5}-?\d{4}", contato):
                                telefone_temp.append(contato)
                            # Verificando se é um email
                            elif "@" in contato:
                                email_temp.append(contato)
                            # Verificando se é um site
                            elif re.search(r"https|\.com|\.br", contato):
                                site_temp.append(contato)
                            # Verificando se é um endereço
                            elif re.search(r"[Ee]ndereço:", contato):
                                endereco_temp.append(re.sub(r"[Ee]ndereço:", "", contato).strip())
                            # Caso contrário, é tratado como observação
                            else:
                                observacao_temp.append(contato)
    
                        # Convertendo listas temporárias em strings separadas por "/"
                        telefone_final.append(" / ".join(telefone_temp) if telefone_temp else "-")
                        email_final.append(" / ".join(email_temp) if email_temp else "-")
                        site_final.append(" / ".join(site_temp) if site_temp else "-")
                        end_final.append(" / ".join(endereco_temp) if endereco_temp else "-")
                        observacao_final.append(" ".join(observacao_temp) if observacao_temp else "-")

                # Definindo a tabela
                tabela_abertura = form[['analista_pesquisador_(solicitante)',
                                        'coletor_escritório_(responsável)',
                                        'data_do_retorno',
                                        'uf_do_escritório',
                                        'job',
                                        'status_do_item',
                                        'elementar',
                                        'item',
                                        'periodicidade',
                                        'uf_do_preço',
                                        'empresa',
                                        'cnpj']].copy()

                tabela_abertura['Data da solicitação'] = form['data_do_mapeamento'].dt.strftime("%d/%m/%Y")
                tabela_abertura['Descrição comercial'] = form['descrição_a_ser_pesquisada']
                tabela_abertura['Endereço'] = end_final
                tabela_abertura['Telefone'] = telefone_final
                tabela_abertura['site'] = site_final
                tabela_abertura['Email'] = email_final
                tabela_abertura['cód. do Formulário de retorno'] = '-'
                tabela_abertura['Observação pesquisador'] = observacao_final

                # Reordenando as colunas
                tabela_abertura = tabela_abertura[['analista_pesquisador_(solicitante)', 'coletor_escritório_(responsável)', 'Data da solicitação', 'data_do_retorno', 'uf_do_escritório',
                                                'job', 'status_do_item', 'elementar', 'item', 'Descrição comercial',
                                                'periodicidade', 'uf_do_preço', 'empresa', 'cnpj',
                                                'Endereço', 'Telefone', 'site', 'Email', 'cód. do Formulário de retorno',
                                                'Observação pesquisador']]
            
                tabela_abertura.reset_index(drop=True, inplace=True)

                planilha = wb_abertura.active

                # Escrevendo dados da Tabela no arquivo Excel
                start_row = 6
                for r, row in tabela_abertura.iterrows():
                    for c, value in enumerate(row, start=1):
                        planilha.cell(row=start_row + r, column=c, value=value)

                # Escrevendo dados da tabela
                #tabela_abertura.to_excel(wb_abertura, sheet_name="Solicitação de novos Inform.", index=False, startrow=5, startcol=0)

                initRow = 6
                endRow = initRow + len(tabela_abertura) - 1

                # Definindo estilos de formatação
                font_body = Font(name='Calibri', size=8)
                alignment_body = Alignment(horizontal='center', vertical='center', wrap_text=False)
                border_body = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                fill_body = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')

                font_item = Font(name='Calibri', size=8)
                alignment_item = Alignment(horizontal='center', vertical='center', wrap_text=True)
                border_item = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                fill_item = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')

                # Adicionando estilos às células
                for row in range(initRow, endRow+1):
                    for col in range(1, 21):
                        cell = planilha.cell(row=row, column=col)
                        cell.font = font_body
                        cell.alignment = alignment_body
                        cell.border = border_body

                # Aplicando estilos - Coluna 9
                for row in range(initRow, endRow + 1):
                    cell = planilha.cell(row=row, column=9)
                    cell.font = font_item
                    cell.alignment = alignment_item
                    cell.border = border_item
                    cell.fill = fill_item

                # Aplicando estilos - Coluna 13
                for row in range(initRow, endRow + 1):
                    cell = planilha.cell(row=row, column=13)
                    cell.font = font_item
                    cell.alignment = alignment_item
                    cell.border = border_item
                    cell.fill = fill_item

                # Aplicando estilos - Colunas 16 a 18
                for row in range(initRow, endRow + 1):
                    for col in range(16, 19):
                        cell = planilha.cell(row=row, column=col)
                        cell.font = font_item
                        cell.alignment = alignment_item
                        cell.border = border_item
                        cell.fill = fill_item

                # Salvando o Workbook
                data = datetime.today().strftime("%d%m%Y")

                wb_abertura.save(f"{temp_dir}/{data}_{solicitante} - Solicitação de Abertura Novos Informantes ({uf_escritorio}) - {job}.xlsx")

    return st.write("Aberturas finalizadas!")
