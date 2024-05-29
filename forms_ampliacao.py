import streamlit as st
import ui as ui
import data_processing 
import pandas as pd
import numpy as np
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles.fills import PatternFill
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import tempfile

def form_ampliacao(content):
    df_ampliacao = content

    # Arrumar nomes das colunas
    df_ampliacao.columns = df_ampliacao.columns.str.strip().str.lower().str.replace(' ', '_')

    # Padronizando a escrita da coluna 'abertura/amplia��o'
    df_ampliacao['abertura_/_amplia��o'] = df_ampliacao['abertura_/_amplia��o'].str.upper()

    # Filtrando a �ltima data e o tipo de formul�rio
    df_filtrado_ampliacao = df_ampliacao[df_ampliacao['abertura_/_amplia��o'] == 'AMPLIA��O']
    latest_date = df_filtrado_ampliacao['data_do_mapeamento'].max()
    df_filtrado_ampliacao = df_filtrado_ampliacao[df_filtrado_ampliacao['data_do_mapeamento'] == latest_date]

    # Removendo o DataFrame original para liberar mem�ria
    del df_ampliacao

    if df_filtrado_ampliacao.empty:
        print("\nN�o h� solicita��o de abertura.")
    else:
        jobs = df_filtrado_ampliacao['job'].unique()
    
        for job in jobs:
            filtro_job = df_filtrado_ampliacao[df_filtrado_ampliacao['job'] == job]
            informantes = filtro_job['bp_(caso_seja_amplia��o)'].unique()
        
            for informante in informantes:
                wb_ampliacao = load_workbook(filename="scr/template_ampliacao.xlsx")
                form = filtro_job[filtro_job['bp_(caso_seja_amplia��o)'] == informante]

                informante = str(informante)[:-2]

                planilha = wb_ampliacao.active

                # Respondente
                resp = form['bp_(caso_seja_amplia��o)'].unique()
                resp = resp[0]
                resp = str(resp)[:-2]
                planilha.cell(row=6, column=3, value=resp)

                # Coletor
                coletor = form['coletor_escrit�rio_(respons�vel)'].unique()
                coletor = coletor[0]
                planilha.cell(row=7, column=3, value=coletor)

                # UF Coletor
                uf_coletor = form['uf_do_escrit�rio'].unique()
                uf_coletor = uf_coletor[0]
                planilha.cell(row=8, column=3, value=uf_coletor)

                # Tipo de Pre�o
                tp_preco = form['tipo_de_pre�o'].unique()
                tp_preco = tp_preco[0]
                planilha.cell(row=9, column=3, value=tp_preco)

                # Periodicidade
                per = form['periodicidade'].unique()
                per = per[0]
                planilha.cell(row=10, column=3, value=per)

                # Solicitante
                sol = form['analista_pesquisador_(solicitante)'].unique()
                sol = sol[0]
                planilha.cell(row=6, column=7, value=sol)

                # Vertical
                vert = form['vertical'].unique()
                vert = vert[0]
                planilha.cell(row=7, column=7, value=vert)

                # Data da Solicita��o
                dt_sol = form['data_do_mapeamento'].dt.strftime("%d/%m/%Y").unique()
                dt_sol = dt_sol[0]
                planilha.cell(row=8, column=7, value=dt_sol)

                # Prazo de Retorno
                retorno = form['prazo_de_retorno'].dt.strftime("%d/%m/%Y").unique()
                retorno = retorno[0]
                planilha.cell(row=9, column=7, value=retorno)

                # Job/Servi�o
                job_serv = form['job'].unique()
                job_serv = job_serv[0]

                # C�digo Item
                cod = form['elementar'].unique()
                cod = cod[0]

                # Definindo a tabela
                tabela_ampliacao = form[['job',
                                         'elementar',
                                         'item',
                                         'unidade_de_medida',
                                         'uf_do_pre�o',
                                         'observa��o_pesquisador_/_informante']].copy()

                tabela_ampliacao['vazio1'] = ''
                tabela_ampliacao['vazio2'] = ''

                # Reordenando as colunas
                tabela_ampliacao = tabela_ampliacao[['job',
                                                     'elementar',
                                                     'item',
                                                     'unidade_de_medida',
                                                     'uf_do_pre�o',
                                                     'vazio1',
                                                     'vazio2',
                                                     'observa��o_pesquisador_/_informante']]
            
                tabela_ampliacao.reset_index(drop=True, inplace=True)

                # Escrevendo dados da Tabela no arquivo Excel
                start_row = 13
                for r, row in tabela_ampliacao.iterrows():
                    for c, value in enumerate(row, start=2):
                        planilha.cell(row=start_row + r, column=c, value=value)

                # Escrevendo dados da tabela
                #tabela_abertura.to_excel(wb_ampliacao, sheet_name="Solicita��o de novos Inform.", index=False, startrow=5, startcol=0)

                initRow = 13
                endRow = initRow + len(tabela_ampliacao) - 1

                # Definindo estilos de formata��o
                font_body = Font(name='Calibri light', size=11)
                alignment_body = Alignment(horizontal='left', vertical='top', wrap_text=False)
                border_body = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                fill_body = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')

                # Adicionando estilos �s c�lulas
                for row in range(initRow, endRow+1):
                    for col in range(2, 10):
                        cell = planilha.cell(row=row, column=col)
                        cell.font = font_body
                        cell.alignment = alignment_body
                        cell.border = border_body

                # Criando uma linha vazia
                start_row = endRow + 1
                for c in range(2, 10):
                    planilha.cell(row=start_row, column=c, value='')
                
                # Criando a caixa com a palavra: "Obs: e aplicando formata��o espec�fica"
                start_row = endRow + 2
                planilha.cell(row=start_row, column=2, value='Obs:')

                # Definindo estilos de formata��o
                font_body = Font(name='Calibri', size=11, bold=True)
                alignment_body = Alignment(horizontal='left', vertical='top', wrap_text=False)
                border_body = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
                fill_body = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')

                # Adicionando estilos �s c�lulas
                for row in range(start_row, start_row + 3):
                    for col in range(2, 10):
                        cell = planilha.cell(row=row, column=col)
                        cell.font = font_body
                        cell.alignment = alignment_body
                        cell.border = border_body

                # Criar o intervalo de c�lulas a serem mescladas
                col_range = openpyxl.utils.get_column_letter(2) + str(start_row)
                col_range += ':' + openpyxl.utils.get_column_letter(9) + str(start_row + 3)
    
                # Mesclar as c�lulas
                planilha.merge_cells(col_range)

                # Salvando o Workbook
                data = datetime.today().strftime("%d%m%Y")

                # Cria um diret�rio tempor�rio
                with tempfile.TemporaryDirectory() as output_dir:
                    # Caminho do arquivo tempor�rio dentro do diret�rio tempor�rio
                    wb_ampliacao.save(f"{output_dir}/Ampliacao BP {informante} - {job} - {coletor}.xlsx")

    return st.write("Amplia��es finalizadas!")